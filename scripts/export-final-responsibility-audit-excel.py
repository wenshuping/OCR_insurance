#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LIGHT_FILL = PatternFill("solid", fgColor="D9EAF7")


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def reason_text(value) -> str:
    if isinstance(value, list):
        return "；".join(str(v) for v in value if v)
    return str(value or "")


def quality_reason(row: dict) -> str:
    return str(row.get("qualityReason") or row.get("qualityReasonText") or "")


def suggestion_for(row: dict) -> str:
    status = row.get("status") or ""
    q_reason = quality_reason(row)
    reasons = reason_text(row.get("reasons"))
    if "missing_product_or_url" in q_reason:
        return "源数据缺产品名或来源链接，需先补齐源数据后再回填。"
    if "pdf_unavailable" in q_reason:
        return "官方链接当前无法下载或文件不可用，需人工确认链接或重新获取条款。"
    if "no_responsibility_text" in q_reason:
        return "已回源解析但未定位到保险责任正文，需人工核对源文件目录或扫描质量。"
    if status == "valid_partial" or "疑似" in reasons or "过短" in reasons:
        return "正文存在但疑似不完整，需人工确认是否为完整保险责任段。"
    return "保留待人工核验。"


def set_sheet_style(ws, widths: dict[int, int] | None = None):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if widths:
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = width


def append_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--audit-json", required=True, type=Path)
    parser.add_argument("--initial-json", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    audit = load_json(args.audit_json)
    initial = load_json(args.initial_json)
    suspects = audit.get("suspects", [])
    initial_count = len(initial) if isinstance(initial, list) else len(initial.get("suspects", []))
    final_count = len(suspects)
    fixed_count = max(initial_count - final_count, 0)
    status_counts = Counter(row.get("status") or "" for row in suspects)

    wb = Workbook()
    ws = wb.active
    ws.title = "剩余问题明细"
    detail_headers = [
        "序号",
        "状态",
        "问题原因",
        "本地ID",
        "飞书记录ID",
        "保险公司",
        "产品名称",
        "产品分类",
        "资料类型",
        "标题",
        "来源链接",
        "质量状态",
        "质量问题",
        "当前摘录",
        "处理建议",
    ]
    detail_rows = []
    for idx, row in enumerate(suspects, 1):
        detail_rows.append(
            [
                idx,
                row.get("status") or "",
                reason_text(row.get("reasons")),
                row.get("localId") or row.get("id") or "",
                row.get("recordId") or "",
                row.get("company") or row.get("feishuTableName") or "",
                row.get("productName") or "",
                row.get("productType") or "",
                row.get("materialType") or "",
                row.get("title") or "",
                row.get("url") or "",
                row.get("qualityStatus") or "",
                quality_reason(row),
                row.get("excerpt") or "",
                suggestion_for(row),
            ]
        )
    append_rows(ws, detail_headers, detail_rows)
    set_sheet_style(
        ws,
        {
            1: 8,
            2: 18,
            3: 28,
            4: 12,
            5: 18,
            6: 16,
            7: 36,
            8: 14,
            9: 12,
            10: 38,
            11: 55,
            12: 22,
            13: 34,
            14: 45,
            15: 34,
        },
    )

    ws_company = wb.create_sheet("公司汇总")
    company_rows = defaultdict(list)
    for row in suspects:
        company_rows[row.get("company") or row.get("feishuTableName") or "未识别"].append(row)
    summary_rows = []
    for company, rows in company_rows.items():
        counter = Counter(r.get("status") or "" for r in rows)
        q_counter = Counter(quality_reason(r) for r in rows if quality_reason(r))
        summary_rows.append(
            [
                company,
                len(rows),
                counter.get("invalid_empty", 0),
                counter.get("valid_partial", 0),
                q_counter.most_common(1)[0][0] if q_counter else "",
            ]
        )
    summary_rows.sort(key=lambda item: (-item[1], item[0]))
    append_rows(ws_company, ["保险公司", "剩余问题数", "invalid_empty", "valid_partial", "主要质量问题"], summary_rows)
    set_sheet_style(ws_company, {1: 20, 2: 14, 3: 16, 4: 16, 5: 45})

    ws_overview = wb.create_sheet("修复汇总")
    overview_rows = [
        ["指标", "值"],
        ["初始待修复问题数", initial_count],
        ["最终剩余问题数", final_count],
        ["已修复/已清除问题数", fixed_count],
        ["清除率", f"{(fixed_count / initial_count * 100):.2f}%" if initial_count else ""],
        ["最终 invalid_empty", status_counts.get("invalid_empty", 0)],
        ["最终 valid_partial", status_counts.get("valid_partial", 0)],
        ["最终审计时间", audit.get("generatedAt") or ""],
        ["最终扫描表数", audit.get("tablesScanned") or ""],
        ["最终扫描行数", audit.get("rowsScanned") or ""],
        ["最终审计 JSON", str(args.audit_json)],
    ]
    for row in overview_rows:
        ws_overview.append(row)
    ws_overview["A1"].fill = HEADER_FILL
    ws_overview["A1"].font = HEADER_FONT
    ws_overview["B1"].fill = HEADER_FILL
    ws_overview["B1"].font = HEADER_FONT
    ws_overview.column_dimensions["A"].width = 24
    ws_overview.column_dimensions["B"].width = 80
    for row in ws_overview.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws_overview["A"]:
        if cell.row > 1:
            cell.fill = LIGHT_FILL

    ws_tables = wb.create_sheet("审计表汇总")
    table_rows = []
    for report in audit.get("tableReports", []):
        counts = report.get("statusCounts") or {}
        table_rows.append(
            [
                report.get("tableName") or "",
                report.get("rows") or 0,
                report.get("problemRows") or 0,
                counts.get("invalid_empty", 0),
                counts.get("valid_partial", 0),
                report.get("tableId") or "",
            ]
        )
    table_rows.sort(key=lambda item: (-item[2], item[0]))
    append_rows(ws_tables, ["表名", "扫描行数", "问题数", "invalid_empty", "valid_partial", "tableId"], table_rows)
    set_sheet_style(ws_tables, {1: 20, 2: 12, 3: 12, 4: 16, 5: 16, 6: 24})

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(args.output)


if __name__ == "__main__":
    main()
